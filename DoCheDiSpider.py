import requests
import re
from lxml import etree
from openpyxl import Workbook
#author：王天琛
# 新建一个数据库
sql_data = []
sql_data_remark = []
for page in range(1, 53):
    # 确定懂车帝的url
    url = 'https://www.dongchedi.com/auto/series/score/4857-x-SO-x-x-x-' + str(page)
    # UA伪装
    headers = {
        'User-Agent': 'User-Agent: Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.102 Safari/537.36 Edg/98.0.1108.56'}
    # 对url主页面抓取
    response = requests.get(url, headers, proxies={'https://': '47.92.234.75'}).text
    response_etree = etree.HTML(response)
    url_lists = response_etree.xpath('//div[@class="tw-flex tw-flex-row-reverse tw-my-2"]/a/@href')
    count = 0
    # print(url_lists)
    for url_list in url_lists:
        #车辆名称
        car_name = response_etree.xpath(
            '//section[@class="tw-col-span-40 tw-pt-16 xl:tw-col-span-28 tw-p-12 xl:tw-pt-0"]/header/h2/span[2]/text()')[
            count].strip()
        # 对评论页中的评论抓取
        url_list1 = 'https://www.dongchedi.com' + url_list
        remark = requests.get(url_list1, headers).text
        remark_text = etree.HTML(remark)
        content = remark_text.xpath('//*[@id="__next"]/div[1]/div/div[2]/div[1]/div[1]/div[2]/div[1]/div[3]/p/span/text()')[
            0]+'【'
        # 购买时间
        ex_buyer = '购车时间.*?</p><p class="jsx-1173095375 bottom-time">(.*?)</p>'
        buyTime = re.findall(ex_buyer, remark, re.S)
        # print(buyTime)
        # 车价
        ex_price = '裸车价.*?</p><p class="jsx-1173095375 bottom-time">(.*?)<!.*?</p>'
        price = re.findall(ex_price, remark, re.S)
        # print(price)
        # 购车地
        ex_location = '购车地.*?</p><p class="jsx-1173095375 bottom-time">(.*?)</p>'
        location = re.findall(ex_location, remark, re.S)
        # print(location)
        # 油耗
        ex_oil_consumption = '百公里油耗.*?</p><p class="jsx-1173095375 bottom-time">(.*?)<!.*?</p>'
        oil_consumption = re.findall(ex_oil_consumption, remark, re.S)
        # print(oil_consumption)
        row1 = []
        row2 = [buyTime, price,location,oil_consumption]
        for f in row2:
            if len(f) == 1:
                f = f[0]
            else:
                f = 'None'
            row1.append(f)
        # 找到最满意的评论，如果没有则为none
        satisfied = re.findall('\【最满意\】[/n]?(.*?)\【', content, re.S)
        # 找到最不满意的评论，如果没有则为none
        dissatisfied = re.findall('\【最不满意\】[/n]?(.*?)\【', content, re.S)
        # 购车经历的评论
        buy_experience = re.findall('\【购车经历\】[/n]?(.*?)\【', content, re.S)
        # 驾驶感受的评论
        driving_feeling = re.findall('\【驾驶感受\】[/n]?(.*?)\【', content, re.S)
        # 乘坐体验的评论
        car_experience = re.findall('\【乘坐体验\】[/n]?(.*?)\【', content, re.S)
        # 保养的评论
        maintenance = re.findall('\【保养\】[/n]?(.*?)\【', content, re.S)
        row3 = []
        row4 = [satisfied, dissatisfied, buy_experience, driving_feeling, car_experience, maintenance]
        for f in row4:
            if len(f) == 1:
                f = f[0]
            else:
                f = 'None'
            row3.append(f)
        #评论页中的外观评价
        appearance = remark_text.xpath('//div[@class="jsx-1173095375 score tw-flex"]/div[1]/p[2]/text()')[0]
        #配置
        configuration = remark_text.xpath('//div[@class="jsx-1173095375 score tw-flex"]/div[2]/p[2]/text()')[0]
        #控制
        control = remark_text.xpath('//div[@class="jsx-1173095375 score tw-flex"]/div[3]/p[2]/text()')[0]
        #内饰
        interior = remark_text.xpath('//div[@class="jsx-1173095375 score tw-flex"]/div[4]/p[2]/text()')[0]
        #空间
        space = remark_text.xpath('//div[@class="jsx-1173095375 score tw-flex"]/div[5]/p[2]/text()')[0]
        #动力
        power = remark_text.xpath('//div[@class="jsx-1173095375 score tw-flex"]/div[6]/p[2]/text()')[0]
        #舒适
        comfort = remark_text.xpath('//div[@class="jsx-1173095375 score tw-flex"]/div[7]/p[2]/text()')[0]
        row5 = [car_name,appearance, configuration, control, interior, space, power,
           comfort]
        row=row5+row1
        count = count + 1
        sql_data.append(row)
        sql_data_remark.append(row3)
# 新建一个excel文件
excel_filename = "DoCheDi"
# 新建一个表
wb = Workbook()
ws1 = wb.active
# 增加表头
title_list = ['车名', '外观', '配置', '控制', '内饰', '空间', '动力', '舒适', '购买时间', '价格', '地址', '油耗']
for row in range(len(title_list)):
    c = row + 1
    ws1.cell(row=1, column=c, value=title_list[row])
# 填写表的内容
for list_index in range(len(sql_data)):
    ws1.append(sql_data[list_index])
# 保存文档
wb.save("./{}.xlsx".format(excel_filename))


#新建一个excel，用于保存评论
excel_filename2 = "pinglun"
# 新建一个表
wb2 = Workbook()
ws2 = wb2.active
# 增加表头
title_list2 = ['最满意','最不满意','购车经历','驾驶感受','乘坐体验','保养']
for row in range(len(title_list2)):
    c = row + 1
    ws2.cell(row=1, column=c, value=title_list2[row])
# 填写表的内容
for list_index in range(len(sql_data_remark)):
    ws2.append(sql_data_remark[list_index])
# 保存文档
wb2.save("./{}.xlsx".format(excel_filename2))
